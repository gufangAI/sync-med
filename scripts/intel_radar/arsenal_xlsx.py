#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""把鹰眼的 GitHub 军火候选渲染成 **Excel 工作簿**(完整版 + 分类分页)。

立此因(创始人 2026-09-02):
  「输出 xls 格式,整整齐齐,看的更好!我需要完整版,你不要搞阉割版」
  「需要分类输出」

所以:**24 个字段一个不落**(md 版为了可读只挑了 6 列,这里全给),
并且**按我们的产线分 sheet**,一页一条线,直接看"这条线能用什么"。

工作簿结构:
  ① 总榜        —— 全部候选,按日均涨星降序,24 列全字段
  ② 优先候选     —— 宽松许可 + 军火库里还没有的新发现(最该先看的)
  ③~⑧ 按产线分页 —— RAG检索 / 古籍OCR / 视频 / 图文 / 方剂知识 / AI免费池 / 未归类
  ⑨ 字段说明     —— 每列什么意思、枚举值有哪些(免得看不懂)

零判断:只客观呈现与排序,不替创始人做采纳决定(与 candidates.json 的 contract 一致)。
数据源:reports/arsenal/candidates.json(鹰眼 arsenal_radar 产出),不重复抓 GitHub。
"""
import io
import json
import os
import sys
import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
ARSENAL = os.path.join(ROOT, "reports", "arsenal")
CAND = os.path.join(ARSENAL, "candidates.json")

LINE_CN = {
    "rag": "RAG检索线", "ocr": "古籍OCR线", "video": "视频产线",
    "image": "图文产线", "formula": "方剂知识线", "modelpool": "AI免费池",
    "unknown": "未归类",
}
FORM_CN = {
    "skill": "技能包", "mcp": "MCP服务", "library": "代码库",
    "architecture": "架构参考", "deploy": "可部署服务", "dataset": "数据集",
    "unknown": "未定",
}
LIC_CN = {"permissive": "宽松·可进闭源", "copyleft": "传染·只能读架构",
          "none": "无许可·不可用", "unknown": "未知·需人工确认"}

def _weekly(c):
    """周涨星数。优先用两次扫描之间的真实增量(stars_delta 换算成 7 天口径);
    没有区间数据时(首次收录的仓)退回 日均涨星 × 7 —— 前者是真实测量,后者是
    生涯均值估算,两者混排会误导,故在"周涨星"旁另有"数据来源"标注。"""
    d, w = c.get("stars_delta"), c.get("stars_delta_window_days")
    if isinstance(d, (int, float)) and isinstance(w, (int, float)) and w > 0:
        return round(d * 7.0 / w, 1)
    spd = c.get("stars_per_day_lifetime")
    return round(spd * 7, 1) if isinstance(spd, (int, float)) else None


def _weekly_src(c):
    d, w = c.get("stars_delta"), c.get("stars_delta_window_days")
    if isinstance(d, (int, float)) and isinstance(w, (int, float)) and w > 0:
        return "实测(%d天区间)" % w
    return "估算(生涯日均×7)"


# 列序(创始人 2026-09-02:「名字,总星数,周涨星数,功能」)——
# 挑东西时眼睛落的顺序就是这个:先认名字,再看多热,再看涨得快不快,最后看能干嘛。
# 其余字段(许可/产线/链接/元数据)按重要性依次往后。
COLS = [
    ("名字",            lambda c: c.get("repo") or ""),
    ("总星数",          lambda c: c.get("stars") or 0),
    ("周涨星数",        _weekly),
    ("功能",            lambda c: c.get("capability") or c.get("description") or ""),
    ("落哪条产线",      lambda c: "/".join(LINE_CN.get(x, x) for x in (c.get("line_candidates") or []))),
    ("可吸收形式",      lambda c: "/".join(FORM_CN.get(x, x) for x in (c.get("transfer_forms") or []))),
    ("许可类型",        lambda c: LIC_CN.get(c.get("license_family") or "unknown", "未知")),
    ("是否新发现",      lambda c: "是" if c.get("new_to_us") else "否"),
    ("链接",            lambda c: c.get("url") or ""),
    ("周涨星来源",      _weekly_src),
    ("日均涨星",        lambda c: c.get("stars_per_day_lifetime")),
    ("原始描述",        lambda c: c.get("description") or ""),
    ("证据",            lambda c: c.get("evidence") or ""),
    ("许可证",          lambda c: c.get("license") or ""),
    ("语言",            lambda c: c.get("lang") or ""),
    ("军火库已有",      lambda c: "是" if c.get("known_in_arsenal") else "否"),
    ("已蒸馏",          lambda c: "是" if c.get("distilled") else "否"),
    ("话题标签",        lambda c: ", ".join(c.get("topics") or [])),
    ("创建日期",        lambda c: c.get("created_at") or ""),
    ("最后推送",        lambda c: c.get("pushed_at") or ""),
    ("仓龄(天)",        lambda c: c.get("age_days")),
    ("命中查询数",      lambda c: c.get("matched_by_n_queries")),
    ("发现维度",        lambda c: ((c.get("found_by") or {}).get("dimension") or "")),
    ("发现查询式",      lambda c: ((c.get("found_by") or {}).get("query") or "")),
]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
NEW_FILL = PatternFill("solid", fgColor="FFF2CC")     # 新发现:淡黄
COPYLEFT_FILL = PatternFill("solid", fgColor="FCE4E4")  # 传染许可:淡红
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(ws, rows, note=None):
    r = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = Font(name="Arial", size=10, italic=True, color="595959")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        r = 2
    head_row = r
    for j, (h, _) in enumerate(COLS, 1):
        c = ws.cell(row=head_row, column=j, value=h)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, cd in enumerate(rows, head_row + 1):
        lic_fam = cd.get("license_family")
        for j, (_, fn) in enumerate(COLS, 1):
            cell = ws.cell(row=i, column=j, value=fn(cd))
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(j in (4, 12, 13, 18, 24)))
            if j == 2:
                cell.number_format = "#,##0"
            if j in (3, 11):
                cell.number_format = "0.0"
        if cd.get("new_to_us"):
            ws.cell(row=i, column=8).fill = NEW_FILL
        if lic_fam == "copyleft":
            ws.cell(row=i, column=7).fill = COPYLEFT_FILL
    # 列宽
    widths = [32, 11, 11, 48, 16, 16, 18, 11, 46, 18, 10, 44, 34, 16, 12, 11, 9, 26, 12, 12, 10, 11, 14, 44]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    ws.auto_filter.ref = "A%d:%s%d" % (head_row, get_column_letter(len(COLS)), head_row + len(rows))


def main():
    if not os.path.isfile(CAND):
        print("arsenal_xlsx: 找不到 %s —— 先让鹰眼跑一轮" % CAND)
        return 1
    d = json.load(io.open(CAND, encoding="utf-8"))
    cands = d.get("candidates") or []
    gen = d.get("generated") or datetime.date.today().isoformat()
    if not cands:
        print("arsenal_xlsx: candidates 为空,不生成空表")
        return 0

    hot = sorted(cands, key=lambda c: (_weekly(c) or 0), reverse=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "总榜"
    write_sheet(ws, hot, "GitHub 军火榜 · %s · 共 %d 个候选 · 按周涨星降序"
                "(总星数偏袒老仓,周涨星才看得出现在热不热)· 数据源 candidates.json"
                % (gen, len(hot)))

    pick = [c for c in hot if c.get("license_family") == "permissive" and c.get("new_to_us")]
    ws2 = wb.create_sheet("优先候选")
    write_sheet(ws2, pick, "宽松许可(可进闭源生产)+ 军火库里还没有的新发现 —— 共 %d 个,"
                           "这是最该先看的一批" % len(pick))

    by_line = {}
    for c in hot:
        for ln in (c.get("line_candidates") or ["unknown"]):
            by_line.setdefault(ln, []).append(c)
    for ln in ["rag", "ocr", "video", "image", "formula", "modelpool", "unknown"]:
        items = by_line.get(ln)
        if not items:
            continue
        w = wb.create_sheet(LINE_CN.get(ln, ln))
        write_sheet(w, items, "可用在【%s】的候选 · 共 %d 个 · 按周涨星降序"
                    % (LINE_CN.get(ln, ln), len(items)))

    # 字段说明页
    wsd = wb.create_sheet("字段说明")
    wsd.column_dimensions["A"].width = 18
    wsd.column_dimensions["B"].width = 96
    docs = [
        ("列名", "含义"),
        ("日均涨星", "总星数 ÷ 仓龄天数。热度的真判据 —— 总星数偏袒老仓,日均涨星才看得出现在热不热。"),
        ("区间涨星", "两次鹰眼扫描之间的星数增量。首次收录的仓为空,跨 run 才积累得出。"),
        ("许可类型", "宽松=MIT/Apache/BSD,可进闭源生产;传染=GPL/AGPL,代码碰了我们的闭源就得开源,"
                     "只能读架构不能抄代码;无许可=默认保留全部权利,法律上不可用。"),
        ("可吸收形式", "技能包/MCP服务/代码库/架构参考/可部署服务/数据集 —— 决定"
                       "「直接装」还是「只读源码吸收」。"),
        ("落哪条产线", "RAG检索线/古籍OCR线/视频产线/图文产线/方剂知识线/AI免费池/未归类。"),
        ("是否新发现", "军火库(arsenal.json)里还没有的 = 是。黄底标出。"),
        ("发现查询式", "鹰眼用哪条 GitHub 搜索式扫到它的 —— 可复现、可审计。"),
        ("", ""),
        ("免责", "本表只做客观呈现与排序,不替你做采纳判断。选哪个是你的决定。"),
    ]
    for i, (a, b) in enumerate(docs, 1):
        ca = wsd.cell(row=i, column=1, value=a)
        cb = wsd.cell(row=i, column=2, value=b)
        ca.font = Font(name="Arial", bold=(i == 1), size=10)
        cb.font = Font(name="Arial", bold=(i == 1), size=10)
        cb.alignment = Alignment(wrap_text=True, vertical="top")

    out = os.path.join(ARSENAL, "军火榜_%s.xlsx" % gen)
    wb.save(out)
    wb.save(os.path.join(ARSENAL, "军火榜_最新.xlsx"))
    print("arsenal_xlsx: 已生成 %s" % out)
    print("            工作表: %s" % " / ".join(wb.sheetnames))
    print("            候选 %d 个 · 字段 %d 列 · 同时写了 军火榜_最新.xlsx" % (len(cands), len(COLS)))
    return 0


if __name__ == "__main__":
    sys.exit(main())
