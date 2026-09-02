#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""军火分类对比表 —— 把挖回来的几千个项目按类别归位,每类给详细说明。

立此因(创始人 2026-09-02):
  「出一个类别对比表格,爬虫搜索类,代码优化skill,科研类,社媒类,
    图片生成,视频生成,社媒输出类,网关管理等等,并有详细的说明!」

为什么需要它:挖掘引擎(arsenal_mine.py)一轮能挖到几千个项目 —— 这解决了
「看不见」的问题,但立刻带来第二个问题:**几千行平铺的表没法挑**。
按类别归位之后,想挑视频工具就只看视频那几十行。

数据源(都是本地 JSON,不重复抓 GitHub):
  reports/arsenal/mined.json       挖掘引擎产出,几千个,带 stars/topics/desc
  reports/arsenal/candidates.json  老雷达产出,带 capability/detail_cn 蒸馏结果

分类判据(category_config.yml,创始人可自己改):
  topics 命中权重 3,description/名字命中权重 1。
  **topics 权重更高是有依据的**:topics 是作者自己打的标签,
  description 是营销文案 —— 实测 crawl4ai 的 topics 为空、描述里写满
  "LLM-friendly",光看描述会把它归进 LLM 类而不是爬虫类。
  所以两者都要看,但可靠的那个权重更大。

详细说明走**内部免费池网关**(红线:零按量计费源),给每类的头部项目写四段:
  做什么 / 怎么用 / 对我们有没有用 / 坑
写过的存进 details.json 复用,不重复烧网关。

产出:
  reports/arsenal/军火分类表_<日期>.xlsx   单表 + 类别列 + 表头筛选器
  reports/arsenal/军火分类表_<日期>.md     同内容的 markdown
  (创始人 2026-09-02 明确要求过「一个表格输出,不要多sheet」——
   分类靠"类别"列 + 筛选器,不靠分页;这样能跨类一起排序比较。)
"""
import io
import json
import os
import sys
import time
import threading
import datetime
import urllib.request

# 用 reconfigure 原地改,**不要**写成 sys.stdout = io.TextIOWrapper(sys.stdout.buffer, ...):
# 本文件被别的脚本 import 时会二次包裹,中间那层 TextIOWrapper 无人引用被 GC、
# __del__ 顺手关掉底层 buffer,之后所有 print 抛 "I/O operation on closed file"。
# (2026-09-02 实测复现;crawl_core/fetch.py 已是这个写法,这里对齐。)
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:                                            # noqa: BLE001
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
ARSENAL = os.path.join(ROOT, "reports", "arsenal")
CFG = os.path.join(HERE, "category_config.yml")

# README 取用 arsenal_enrich.gh_readme(它内部走 crawl_core 的取页 + 正文两层)。
# 守着 import 而不是让它炸:这个链条底下压着 lxml / trafilatura 两个非 stdlib 依赖,
# 缺一个就会把**整张军火分类表**炸掉,而 workflow 里这一步是 `|| true`,炸了还看不见。
# 拿不到就退回"只有元数据"的老提示词 —— 最坏等于改造前,但必须出声。
try:
    sys.path.insert(0, HERE)
    from arsenal_enrich import gh_readme as _gh_readme            # noqa: E402
except BaseException as e:                                        # noqa: BLE001
    _gh_readme = None
    print("  [warn] 取不到 arsenal_enrich.gh_readme(%s):详解将只依据元数据,"
          "不读 README —— 详细程度会明显下降" % str(e)[:100])

GATEWAY = os.environ.get("GATEWAY_URL", "https://gufangai.com/api/gateway/chat")
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 " \
     "(KHTML, like Gecko) Chrome/122.0 Safari/537.36"

LIC_CN = {
    "MIT": "MIT · 可进闭源",
    "Apache-2.0": "Apache · 可进闭源",
    "BSD-3-Clause": "BSD · 可进闭源",
    "BSD-2-Clause": "BSD · 可进闭源",
    "ISC": "ISC · 可进闭源",
    "MPL-2.0": "MPL · 文件级传染",
    "GPL-3.0": "GPL · 学思路自研,不抄码",
    "GPL-2.0": "GPL · 学思路自研,不抄码",
    "AGPL-3.0": "AGPL · 学思路自研,不抄码",
    "NOASSERTION": "自定义条款 · 必须人工读",
    "": "无许可 · 默认保留全部权利,不可用",
}


def load_cfg():
    try:
        import yaml
    except ImportError:
        raise SystemExit("缺 pyyaml:pip install pyyaml")
    if not os.path.isfile(CFG):
        raise SystemExit("找不到分类配置 %s" % CFG)
    with io.open(CFG, encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def load_pool():
    """合并两个数据源。同一个仓两边都有时,以挖掘引擎的为准(它的星数是本轮实测),
    但把老雷达蒸馏出的 capability/detail_cn 带过来 —— 那是花过网关调用的成果,
    丢掉等于重复烧一次。"""
    pool = {}
    p = os.path.join(ARSENAL, "mined.json")
    if os.path.isfile(p):
        d = json.load(io.open(p, encoding="utf-8"))
        for r in (d.get("mined") or []):
            if r.get("repo"):
                pool[r["repo"]] = dict(r)
    p = os.path.join(ARSENAL, "candidates.json")
    if os.path.isfile(p):
        d = json.load(io.open(p, encoding="utf-8"))
        for c in (d.get("candidates") or []):
            k = c.get("repo")
            if not k:
                continue
            row = pool.setdefault(k, {
                "repo": k, "stars": c.get("stars"),
                "lang": c.get("lang"), "topics": c.get("topics") or [],
                "desc": c.get("description") or "", "license": c.get("license") or "",
            })
            for f in ("capability", "detail_cn", "evidence"):
                if c.get(f) and not row.get(f):
                    row[f] = c[f]
    return pool


def classify(row, cats):
    """给一个项目定类别。返回 (key, name, 命中分, 命中了哪些词)。

    topics 权重 3、文本权重 1 —— 见文件头说明,这是实测定的不是拍脑袋。
    一个项目可能同时像两类(video-use 既是视频又是浏览器自动化),
    取分最高的那类;并列时按 categories 里的先后 —— 配置顺序即优先级,
    所以配置里把更贴我们产线的类别写在前面。
    """
    topics = {str(t).lower() for t in (row.get("topics") or [])}
    text = " ".join([
        str(row.get("repo") or "").lower(),
        str(row.get("desc") or "").lower(),
        str(row.get("capability") or "").lower(),
    ])
    best = None
    for c in cats:
        hits, score = [], 0
        for t in (c.get("topics") or []):
            if str(t).lower() in topics:
                score += 3
                hits.append(t)
        for w in (c.get("words") or []):
            if str(w).lower() in text:
                score += 1
                hits.append(w)
        if score and (best is None or score > best[2]):
            best = (c["key"], c["name"], score, hits[:6])
    return best


def ask_gateway(prompt, system, timeout=120):
    """走内部免费池网关。红线:严禁任何按量计费源。

    返回契约:成功返回文本;失败返回 "" —— 调用方据此跳过,不写半截内容。
    """
    body = json.dumps({
        "messages": [{"role": "system", "content": system},
                     {"role": "user", "content": prompt}],
        "max_tokens": 900, "temperature": 0.3, "source": "arsenal_category",
    }).encode()
    # X-Gateway-Key 是必需的:内部免费池网关不是公开 API,没这个头一律 UNAUTHORIZED。
    # (2026-09-02 实测踩过:漏了这一行 → 40 个详解全部"网关失败",
    #  而失败信息只写"跳过",不看 curl 根本不知道是缺 key 还是模型不通。)
    # key 走环境变量/Secrets 注入,**绝不写进代码** —— 这个仓是 PUBLIC。
    hdrs = {"Content-Type": "application/json", "User-Agent": UA}
    gk = os.environ.get("GW_KEY", "").strip()
    if gk:
        hdrs["X-Gateway-Key"] = gk
    req = urllib.request.Request(GATEWAY, data=body, method="POST", headers=hdrs)
    try:
        j = json.loads(urllib.request.urlopen(req, timeout=timeout).read())
        # 内部网关返回 {"ok":true,"text":...},不是 OpenAI 的 choices —— 两种都兼容
        t = j.get("text") or ""
        if not t:
            t = ((j.get("choices") or [{}])[0].get("message") or {}).get("content") or ""
        return t.strip()
    except Exception:                                            # noqa: BLE001
        return ""


PLATFORM = """古方 AI 星图 = 中医古籍 AI 平台。现有产线:
1) RAG检索:2100+医案 + 7700+古籍向量检索,讯飞embedding(绑死不可换)
2) 古籍OCR:几万本扫描书页,RapidOCR 跑 GitHub Actions,竖排/夹注识别差
3) 视频产线:Remotion+edge-tts 出专家口播视频,但发布分发=0代码
4) 图文产线:satori 渲方剂卡,缺自动发布
5) AI免费池:自建网关,9家免费模型容错链,零按量计费
6) 鹰眼情报:自建挖掘引擎,五条矿脉自动发现开源
硬约束:无服务器(CF Workers/GitHub Actions)、本地禁算力、AI 只走内部免费池、
绝不换 embedding 供应商、GPL/AGPL 代码不能进闭源生产。"""

SYS = """你是古方AI星图平台的技术选型顾问。给你一个 GitHub 项目,
写一段**详细的中文说明**让平台负责人判断值不值得用。

四段,每段用「」标题开头,段间换行:
「做什么」具体功能点 2-4 个,不要一句概括
「怎么用」装法/调法(pip? docker? 要服务器吗?)
「对我们有没有用」结合平台现状明确说:用在哪条产线的哪一步,或直说"用不上,因为..."
「坑」前提与风险(要GPU? 要付费API? 许可传染? 依赖重?)

规则:总长 150-300 字;只依据给你的信息写,不确定的写"信息不足",**绝不编造**;
该说"用不上"就说"用不上"。"""


def enrich(rows, cache, limit, workers=6):
    """给还没有详解的项目补一段中文详解。已有的直接复用,不重复烧网关。

    **并发跑**。串行版本实测跑 20 分钟一条都没出来:
      · timeout 设了 120 秒 —— 而网关实测只要 1-6 秒,120 秒纯粹是给自己挖坑,
        一个慢调用就能把整批拖住
      · print 到重定向文件时被缓冲,进度看不见,分不清"在跑"还是"卡死"
    两条一起改:超时压到 45 秒,输出强制 flush,再并发 6 路。
    6 路是保守取值 —— 免费池本身能扛更多,但我们没必要为了快去压别人的额度。
    """
    todo = [r for r in rows if not r.get("detail_cn")
            and r["repo"] not in cache][:limit]
    if not todo:
        print("  详解:都已有,跳过", flush=True)
        return 0

    ways = {}
    ways_lock = threading.Lock()

    def one(r):
        # README 才是"详细"的来源。创始人 2026-09-02:「功能介绍详细一点,要不然容易被忽略」
        # —— 而在此之前这个提示词里**只有一句 300 字的 description**,模型再强也变不出
        # 它没见过的信息。README 走 arsenal_enrich.gh_readme(crawl_core 的取页 + 正文两层),
        # **不在这里写第二份**(平台铁律:同一份逻辑只许有一份实现)。
        rm, way = "", "skip"
        if _gh_readme is not None:
            try:
                rm, way = _gh_readme(r["repo"], limit=2500)
            except Exception as e:                               # noqa: BLE001
                rm, way = "", "err:" + type(e).__name__
        with ways_lock:
            ways[way] = ways.get(way, 0) + 1
        p = ("%s\n\n项目:%s(%s★,%s,许可 %s)\ntopics:%s\n描述:%s\n能力:%s"
             % (PLATFORM, r["repo"], r.get("stars"), r.get("lang") or "未知",
                r.get("license") or "未标", ", ".join(r.get("topics") or [])[:200],
                (r.get("desc") or "")[:300], (r.get("capability") or "")[:200]))
        if rm:
            p += "\n\nREADME:\n" + rm
        # 抓不到 README 时**不跳过**,退回原来的"只有元数据"提示词 ——
        # 最坏情况等于改造前的行为,不会因为多了一层而少产出。
        return r, ask_gateway(p, SYS, timeout=45)

    import concurrent.futures as cf
    ok = fail = 0
    t0 = time.time()
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        for i, (r, out) in enumerate(ex.map(one, todo), 1):
            if out:
                cache[r["repo"]] = out
                r["detail_cn"] = out
                ok += 1
                print("  [%d/%d] %s ✓ %d字" % (i, len(todo), r["repo"][:38], len(out)),
                      flush=True)
            else:
                fail += 1
                print("  [%d/%d] %s 网关没返回" % (i, len(todo), r["repo"][:38]),
                      flush=True)
    print("  详解:成功 %d · 失败 %d · 耗时 %.0f 秒"
          % (ok, fail, time.time() - t0), flush=True)
    # 哨兵:README 各条路各走了几次。api 那一路整批变 0 = token 没注进来或被限流,
    # 一眼看得见;否则"详解又变短了"这种退化只能靠人肉察觉。
    print("  README 取法分布:%s"
          % ("  ".join("%s=%d" % kv for kv in sorted(ways.items())) or "(无)"), flush=True)
    return ok


def write_xlsx(groups, cats, out, gen):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLS = [
        ("类别", 16), ("名字", 34), ("总星数", 11), ("功能详解", 96),
        ("语言", 12), ("许可(能不能抄)", 24), ("落哪条产线", 20),
        ("命中依据", 26), ("发现来源", 22), ("链接", 44), ("最后更新", 12),
    ]
    HEAD_FILL = PatternFill("solid", fgColor="1F3864")
    HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY = Font(name="Arial", size=10)
    CAT_FILL = PatternFill("solid", fgColor="DDEBF7")
    RISK_FILL = PatternFill("solid", fgColor="FCE4E4")
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    wb = Workbook()
    ws = wb.active
    ws.title = "军火分类表"

    note = ("GitHub 军火分类表 · %s · 共 %d 类 %d 个项目 · "
            "分类看「类别」列,点表头筛选器过滤 · "
            "红底=传染许可:读架构学思路、自己重写实现,不复制代码 · "
            "星数为实测值,非榜单转抄"
            % (gen, len(groups), sum(len(v) for v in groups.values())))
    ws.cell(row=1, column=1, value=note).font = Font(
        name="Arial", size=10, italic=True, color="595959")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))

    for j, (h, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=j, value=h)
        c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cat_desc = {c["key"]: c.get("desc", "") for c in cats}
    cat_name = {c["key"]: c["name"] for c in cats}
    r = 3
    for ckey in list(groups.keys()):
        rows = groups[ckey]
        if not rows:
            continue
        # 类别分隔行:把这一类"是什么、对我们意味着什么"写在头上,
        # 免得看表的人得回去翻配置才知道这类是干嘛的
        cell = ws.cell(row=r, column=1,
                       value="【%s】%s" % (cat_name.get(ckey, ckey), cat_desc.get(ckey, "")))
        cell.font = Font(name="Arial", bold=True, size=10, color="1F3864")
        cell.fill = CAT_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        ws.row_dimensions[r].height = 30
        r += 1
        for d in rows:
            lic = d.get("license") or ""
            vals = [
                cat_name.get(ckey, ckey),
                d.get("repo"),
                d.get("stars") or 0,
                d.get("detail_cn") or d.get("capability") or d.get("desc") or "(暂无详解)",
                d.get("lang") or "",
                LIC_CN.get(lic, lic or "未知"),
                "/".join(d.get("line_candidates") or []) or "",
                ", ".join(d.get("_hits") or [])[:120],
                (d.get("origin") or [""])[0] if d.get("origin") else "",
                "https://github.com/" + str(d.get("repo") or ""),
                (d.get("pushed_at") or "")[:10],
            ]
            for j, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = BODY
                c.border = BORDER
                c.alignment = Alignment(vertical="top", wrap_text=(j in (4, 8)))
                if j == 3:
                    c.number_format = "#,##0"
            if any(x in lic for x in ("GPL", "AGPL")) or not lic:
                ws.cell(row=r, column=6).fill = RISK_FILL
            ws.row_dimensions[r].height = 120
            r += 1

    for j, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=3, column=1)
    ws.auto_filter.ref = "A2:%s%d" % (get_column_letter(len(COLS)), r - 1)
    wb.save(out)
    return r - 3


def write_md(groups, cats, out, gen):
    cat_desc = {c["key"]: c.get("desc", "") for c in cats}
    cat_name = {c["key"]: c["name"] for c in cats}
    L = ["# 军火分类表 · %s" % gen, "",
         "> 挖掘引擎产出按类别归位。星数为**实测值**,不转抄榜单。",
         "> 许可栏标「传染」的:**读它的架构、学它的做法,再糅合几家自己重写一个** —— "
         "不复制代码即可,不是不能用。", ""]
    L.append("## 各类概览")
    L.append("")
    L.append("| 类别 | 数量 | 这一类是干什么的 |")
    L.append("|---|---:|---|")
    for k, rows in groups.items():
        if rows:
            L.append("| **%s** | %d | %s |"
                     % (cat_name.get(k, k), len(rows), cat_desc.get(k, "")))
    L.append("")
    for k, rows in groups.items():
        if not rows:
            continue
        L.append("## %s(%d)" % (cat_name.get(k, k), len(rows)))
        L.append("")
        L.append("> %s" % cat_desc.get(k, ""))
        L.append("")
        for d in rows:
            lic = d.get("license") or ""
            L.append("### [%s](https://github.com/%s) ★%s"
                     % (d.get("repo"), d.get("repo"), "{:,}".format(d.get("stars") or 0)))
            L.append("")
            L.append("**许可**:%s ｜ **语言**:%s ｜ **发现来源**:%s"
                     % (LIC_CN.get(lic, lic or "未知"), d.get("lang") or "未知",
                        (d.get("origin") or [""])[0] if d.get("origin") else "—"))
            L.append("")
            det = d.get("detail_cn") or d.get("capability") or d.get("desc") or "(暂无详解)"
            L.append(det)
            L.append("")
    io.open(out, "w", encoding="utf-8").write("\n".join(L) + "\n")


def main():
    cfg = load_cfg()
    cats = cfg.get("categories") or []
    if not cats:
        raise SystemExit("分类配置里没有 categories")
    pool = load_pool()
    print("军火分类表 · 候选池 %d 个项目 · %d 个类别" % (len(pool), len(cats)))
    if not pool:
        raise SystemExit("候选池是空的 —— 先让挖掘引擎跑一轮")

    groups = {c["key"]: [] for c in cats}
    unc = []
    for row in pool.values():
        hit = classify(row, cats)
        if hit:
            row["_hits"] = hit[3]
            row["_cscore"] = hit[2]
            groups[hit[0]].append(row)
        else:
            unc.append(row)

    lim = int(cfg.get("per_category_limit", 40) or 0)
    for k in groups:
        # 类内排序:先看分类命中强度,再看星数。**不是纯按星数** ——
        # 纯按星数会让每一类的头部都被通用大项目占满(vscode/react 这种),
        # 而我们要的是"这一类里最对口的",不是"这一类里最有名的"。
        groups[k].sort(key=lambda x: (-(x.get("_cscore") or 0), -(x.get("stars") or 0)))
        if lim:
            groups[k] = groups[k][:lim]

    print("\n分类结果:")
    for c in cats:
        print("  %-14s %4d 个" % (c["name"], len(groups[c["key"]])))
    print("  %-14s %4d 个(分类器认不出,保留不丢弃)"
          % (cfg.get("uncategorized_name", "未分类"), len(unc)))

    # 详解
    cachep = os.path.join(ARSENAL, "details.json")
    cache = {}
    if os.path.isfile(cachep):
        try:
            cache = json.load(io.open(cachep, encoding="utf-8"))
        except Exception:                                        # noqa: BLE001
            cache = {}
    for k in groups:
        for r in groups[k]:
            if not r.get("detail_cn") and cache.get(r["repo"]):
                r["detail_cn"] = cache[r["repo"]]

    per = int(cfg.get("enrich_per_run", 24))
    if per > 0 and os.environ.get("SKIP_ENRICH") != "1":
        print("\n补详解(走内部免费池网关,每类取头部):")
        head = []
        for k in groups:
            head.extend(groups[k][:4])
        n = enrich(head, cache, per)
        if n:
            json.dump(cache, io.open(cachep, "w", encoding="utf-8"),
                      ensure_ascii=False, indent=1)
            print("  新写 %d 条详解,已存 details.json 供下次复用" % n)

    gen = datetime.date.today().isoformat()
    if not os.path.isdir(ARSENAL):
        os.makedirs(ARSENAL)
    xp = os.path.join(ARSENAL, "军火分类表_%s.xlsx" % gen)
    mp = os.path.join(ARSENAL, "军火分类表_%s.md" % gen)
    try:
        n = write_xlsx(groups, cats, xp, gen)
        write_xlsx(groups, cats, os.path.join(ARSENAL, "军火分类表_最新.xlsx"), gen)
        print("\n已生成 %s(%d 行)" % (xp, n))
    except ImportError:
        print("\n缺 openpyxl,跳过 xlsx")
    write_md(groups, cats, mp, gen)
    write_md(groups, cats, os.path.join(ARSENAL, "军火分类表_最新.md"), gen)
    print("已生成 %s" % mp)
    return 0


if __name__ == "__main__":
    sys.exit(main())
